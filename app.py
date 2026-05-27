import streamlit as st
import pandas as pd
import plotly.express as px
import datetime
import io
import unicodedata

# ==========================================
# 1. CẤU HÌNH GIAO DIỆN LIGHT MODE & FONT MONTSERRAT
# ==========================================
st.set_page_config(page_title="SGM Quản trị kho", page_icon="🏢", layout="wide", initial_sidebar_state="expanded")

st.markdown("""
    <style>
    /* Nhúng font Montserrat đồng bộ toàn diện hệ thống */
    @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;500;600;700;800;900&display=swap');
    
    /* CHỈ ÁP DỤNG FONT CHO CÁC THẺ VĂN BẢN (KHÔNG ĐÈ LÊN THƯ VIỆN ICON CỦA GLIDE TABLE VÀ SIDEBAR) */
    html, body, [data-testid="stAppViewContainer"] {
        font-family: 'Montserrat', sans-serif;
    }
    
    .stApp p, .stApp label, .stApp h1, .stApp h2, .stApp h3, .stApp h4, .stApp h5, .stApp h6, .stApp td, .stApp th, .stApp li {
        font-family: 'Montserrat', sans-serif !important;
    }
    
    /* Loại trừ thẻ span thông thường để không phá vỡ icon ghim/ẩn cột của bảng */
    .stApp p, .stApp label, .stApp h1, .stApp h2, .stApp h3, .stApp h4, .stApp h5, .stApp h6 {
        font-family: 'Montserrat', sans-serif !important;
    }

    /* ĐẢM BẢO TUYỆT ĐỐI KHÔNG GHI ĐÈ ICON THU NHỎ/MỞ RỘNG SIDEBAR VÀ HEADER TRÊN MOBILE */
    button[data-testid*="CollapseButton"] *,
    button[class*="CollapseButton"] *,
    button[data-testid="baseButton-header"] *,
    [data-testid="stHeader"] *,
    [class*="Icon"] *,
    .material-icons {
        font-family: inherit !important;
    }

    /* -----------------------------------
       1. KHU VỰC MAIN & SIDEBAR
       ----------------------------------- */
    .stApp, .main { background-color: #ffffff !important; }
    .stApp p, .stApp span, div[data-testid="stMarkdownContainer"] { color: #1e293b !important; }
    h1, h2, h3, h4, h5, h6 { color: #0f172a !important; font-weight: 800 !important; }

    /* Sidebar nền xám nhẹ */
    [data-testid="stSidebar"] {
        background-color: #f1f5f9 !important; 
        border-right: 1px solid #e2e8f0 !important;
    }
    [data-testid="stSidebar"] p, [data-testid="stSidebar"] label { 
        color: #334155 !important; 
        font-weight: 600 !important;
    }
    [data-testid="stSidebar"] h4 { 
        color: #0f172a !important; 
        font-size: 22px !important; 
        font-weight: 900 !important; 
        margin-top: 25px !important;
        margin-bottom: 15px !important;
        border-bottom: 2px solid #cbd5e1;
        padding-bottom: 10px;
    }
    .copyright-text { font-size: 11px !important; color: #64748b !important; text-align: center; margin-top: -15px; margin-bottom: 25px; font-weight: 500; }

    /* Loại bỏ phông nền trắng của logo SGM */
    [data-testid="stSidebar"] img {
        mix-blend-mode: multiply;
    }

    /* Slider & Multiselect với tone màu Green Pastel */
    .stSlider > div > div > div > div { background-color: #cbd5e1 !important; }
    .stSlider > div > div > div > div > div { background-color: #86efac !important; }
    .stSlider > div > div > div > div > div[role="slider"] { background-color: #22c55e !important; box-shadow: 0 0 10px rgba(34,197,94,0.4) !important; }
    
    .stMultiSelect div[data-baseweb="select"] {
        background-color: #f0fdf4 !important; 
        border: 1px solid #86efac !important;
        border-radius: 8px !important;
    }
    .stMultiSelect div[data-baseweb="select"]:focus-within {
        border-color: #22c55e !important;
    }
    .stMultiSelect span[data-baseweb="tag"] { 
        background-color: #bbf7d0 !important; 
        border: 1px solid #86efac !important; 
    }
    .stMultiSelect span[data-baseweb="tag"] span { 
        color: #14532d !important; 
        font-weight: 700 !important; 
    }
    .stMultiSelect span[data-baseweb="tag"] svg {
        fill: #14532d !important;
    }

    /* -----------------------------------
       2. SCALE LẠI HIỂN THỊ CÁC THẺ CARD LỚN GỌN GÀNG, THẨM MỸ HƠN
       ----------------------------------- */
    div[data-testid="stMetric"] {
        background: linear-gradient(145deg, #fffdfa, #fdf4e7) !important; 
        border-radius: 16px !important;
        padding: 16px 20px !important;
        box-shadow: 6px 10px 20px rgba(139,92,26,0.06), -2px -2px 8px rgba(255,255,255,0.8) !important;
        border-left: 5px solid #fb923c !important; 
        border-top: 1px solid rgba(251,146,60,0.1) !important;
        transition: all 0.3s cubic-bezier(0.25, 0.8, 0.25, 1) !important;
    }
    div[data-testid="stMetric"]:hover { 
        transform: translateY(-5px) scale(1.01); 
        box-shadow: 8px 14px 24px rgba(139,92,26,0.1), -2px -2px 10px rgba(255,255,255,0.9) !important; 
    }
    div[data-testid="stMetric"] label { 
        color: #c2410c !important; 
        font-size: 11px !important; 
        text-transform: uppercase; 
        font-weight: 800 !important; 
        letter-spacing: 0.8px; 
    }
    div[data-testid="stMetric"] div[data-testid="stMetricValue"] { 
        color: #1e293b !important; 
        font-size: 26px !important; 
        font-weight: 900 !important; 
        text-shadow: none !important;
        margin-top: 4px;
    }

    /* THẺ METRIC TÙY BIẾN CHO HOVER TOOLTIP TRONG TAB 5 */
    .custom-metric-card {
        background: linear-gradient(145deg, #fffdfa, #fdf4e7) !important; 
        border-radius: 16px !important;
        padding: 16px 20px !important;
        box-shadow: 6px 10px 20px rgba(139,92,26,0.06), -2px -2px 8px rgba(255,255,255,0.8) !important;
        border-left: 5px solid #fb923c !important; 
        border-top: 1px solid rgba(251,146,60,0.1) !important;
        transition: all 0.3s cubic-bezier(0.25, 0.8, 0.25, 1) !important;
        position: relative;
        cursor: pointer;
    }
    .custom-metric-card:hover { 
        transform: translateY(-5px) scale(1.01); 
        box-shadow: 8px 14px 24px rgba(139,92,26,0.1), -2px -2px 10px rgba(255,255,255,0.9) !important; 
    }
    .custom-metric-card .metric-label { 
        color: #c2410c !important; 
        font-size: 11px !important; 
        text-transform: uppercase; 
        font-weight: 800 !important; 
        letter-spacing: 0.8px; 
    }
    .custom-metric-card .metric-value { 
        color: #1e293b !important; 
        font-size: 26px !important; 
        font-weight: 900 !important; 
        margin-top: 4px;
    }
    .custom-metric-card .tooltip-text {
        visibility: hidden;
        width: 280px;
        background-color: #1e293b;
        color: #f8fafc;
        text-align: left;
        border-radius: 12px;
        padding: 15px;
        position: absolute;
        z-index: 9999;
        bottom: 110%;
        left: 50%;
        margin-left: -140px;
        opacity: 0;
        transition: opacity 0.2s, visibility 0.2s;
        box-shadow: 0 10px 25px rgba(0,0,0,0.25);
        font-size: 12px;
        font-weight: 500;
        line-height: 1.5;
        max-height: 250px;
        overflow-y: auto;
        border: 1px solid #475569;
    }
    .custom-metric-card:hover .tooltip-text {
        visibility: visible;
        opacity: 1;
    }

    /* THẺ METRIC ĐƠN HÀNG TRUNG BÌNH CÓ HOVER TOOLTIP TRONG TAB 2 */
    .custom-metric-card-aov {
        background: linear-gradient(145deg, #f8fafc, #f1f5f9) !important; 
        border: 1px solid #cbd5e1 !important; 
        border-left: 5px solid #f59e0b !important; 
        padding: 16px 20px !important; 
        border-radius: 12px; 
        min-height: 100px;
        position: relative;
        cursor: pointer;
    }
    .custom-metric-card-aov .tooltip-text-aov {
        visibility: hidden;
        width: 255px;
        background-color: #1e293b;
        color: #f8fafc;
        text-align: left;
        border-radius: 12px;
        padding: 15px;
        position: absolute;
        z-index: 9999;
        bottom: 110%;
        left: 50%;
        margin-left: -127px;
        opacity: 0;
        transition: opacity 0.2s, visibility 0.2s;
        box-shadow: 0 10px 25px rgba(0,0,0,0.25);
        font-size: 12px;
        font-weight: 500;
        line-height: 1.5;
        max-height: 200px;
        overflow-y: auto;
        border: 1px solid #475569;
    }
    .custom-metric-card-aov:hover .tooltip-text-aov {
        visibility: visible;
        opacity: 1;
    }

    /* IN ĐẬM RÕ NÉT TOÀN BỘ TIÊU ĐỀ (HEADER) CỦA BẢNG DATAFRAME */
    [data-testid="stDataFrame"] th, 
    [data-testid="stDataFrame"] [role="columnheader"] p,
    [data-testid="stDataFrame"] [role="columnheader"] span,
    [data-testid="stDataFrame"] [role="columnheader"] {
        font-weight: 900 !important;
        color: #0f172a !important;
    }

    /* STICKY HERO SECTION TỰ ĐỘNG TRÊN NỀN SÁNG */
    div[data-testid="stVerticalBlock"] > div:has(div[data-testid="stMetric"]) {
        position: sticky;
        top: 2.8rem;
        z-index: 999;
        background-color: #ffffff;
        padding: 10px 0 20px 0;
        border-bottom: 1px solid #e2e8f0;
    }

    /* -----------------------------------
       3. TÙY BIẾN ĐÓNG FRAME CHO CÁC TABS THEO YÊU CẦU
       ----------------------------------- */
    button[data-baseweb="tab"] { 
        font-family: 'Montserrat', sans-serif !important;
        font-size: 19px !important; /* Tăng nhẹ kích cỡ */
        font-weight: 800 !important; /* In đậm */
        color: #475569 !important; /* Màu xám đậm */
        background-color: #f8fafc !important; 
        border: 1px solid #cbd5e1 !important; /* Đóng frame */
        border-radius: 8px !important; 
        margin-right: 10px !important;
        padding: 12px 24px !important;
        transition: all 0.2s ease-in-out !important;
    }
    button[aria-selected="true"] { 
        color: #388e3c !important; /* Màu green của logo */
        background-color: #ffffff !important; 
        border: 2px solid #388e3c !important; /* Khung dày màu green */
        box-shadow: 0 4px 12px rgba(56,142,60,0.1) !important;
    }

    .stDataFrame { 
        background: #ffffff !important; 
        border-radius: 16px; 
        padding: 10px; 
        box-shadow: 0 10px 25px rgba(0,0,0,0.05) !important; 
        border: 1px solid #e2e8f0; 
    }
    
    .stDownloadButton button { 
        font-family: 'Montserrat', sans-serif !important; 
        background-color: #388e3c !important; 
        color: white !important; 
        border-radius: 10px !important; 
        font-weight: 800 !important; 
        border: none !important; 
        width: 100%; 
        box-shadow: 0 8px 15px rgba(56,142,60,0.2); 
        transition: all 0.3s ease; 
    }
    
    /* -----------------------------------
       4. SMART CARDS PHONG CÁCH LIGHT MODE
       ----------------------------------- */
    .smart-card-info { background: #eff6ff !important; border-left: 5px solid #3b82f6 !important; padding: 22px; border-radius: 12px; color: #1e3a8a !important; margin-bottom: 18px; box-shadow: 0 4px 12px rgba(59,130,246,0.08); border-top: 1px solid rgba(59,130,246,0.1); }
    .smart-card-success { background: #f0fdf4 !important; border-left: 5px solid #10b981 !important; padding: 22px; border-radius: 12px; color: #14532d !important; margin-bottom: 18px; box-shadow: 0 4px 12px rgba(16,185,129,0.08); border-top: 1px solid rgba(16,185,129,0.1); }
    .smart-card-error { background: #fef2f2 !important; border-left: 5px solid #ef4444 !important; padding: 22px; border-radius: 12px; color: #7f1d1d !important; margin-bottom: 18px; box-shadow: 0 4px 12px rgba(239,68,68,0.08); border-top: 1px solid rgba(239,68,68,0.1); }
    .smart-card-warning { background: #fffbeb !important; border-left: 5px solid #f59e0b !important; padding: 22px; border-radius: 12px; color: #78350f !important; margin-bottom: 18px; box-shadow: 0 4px 12px rgba(245,158,11,0.08); border-top: 1px solid rgba(245,158,11,0.1); }

    /* THẺ HERO CARD TÙY BIẾN CHO S2S BÌNH QUÂN CÓ HOVER TOOLTIP (LUÔN BAY LÊN TRÊN CARD) */
    .custom-hero-card {
        background: linear-gradient(145deg, #fffdfa, #fdf4e7) !important; 
        border-radius: 16px !important;
        padding: 16px 20px !important;
        box-shadow: 6px 10px 20px rgba(139,92,26,0.06), -2px -2px 8px rgba(255,255,255,0.8) !important;
        border-left: 5px solid #fb923c !important; 
        border-top: 1px solid rgba(251,146,60,0.1) !important;
        transition: all 0.3s cubic-bezier(0.25, 0.8, 0.25, 1) !important;
        position: relative;
        cursor: pointer;
    }
    .custom-hero-card:hover {
        transform: translateY(-5px) scale(1.01); 
        box-shadow: 8px 14px 24px rgba(139,92,26,0.1), -2px -2px 10px rgba(255,255,255,0.9) !important;
    }
    .custom-hero-card .hero-label {
        color: #c2410c !important; 
        font-size: 11px !important; 
        text-transform: uppercase; 
        font-weight: 800 !important; 
        letter-spacing: 0.8px; 
    }
    .custom-hero-card .hero-value {
        color: #1e293b !important; 
        font-size: 26px !important; 
        font-weight: 900 !important; 
        margin-top: 4px;
    }
    .custom-hero-card .tooltip-hero {
        visibility: hidden;
        width: 300px;
        background-color: #1e293b;
        color: #f8fafc;
        text-align: left;
        border-radius: 12px;
        padding: 15px;
        position: absolute;
        z-index: 99999 !important; /* Đẩy lớp hiển thị lên cao nhất để không bị đè */
        bottom: 115%; /* Chuyển thành hiển thị bay lên trên Card */
        left: 50%;
        margin-left: -150px;
        opacity: 0;
        transition: opacity 0.2s, visibility 0.2s;
        box-shadow: 0 -10px 25px rgba(0,0,0,0.25); /* Bóng đổ hướng lên trên */
        font-size: 12px;
        font-weight: 500;
        line-height: 1.5;
        border: 1px solid #475569;
    }
    .custom-hero-card:hover .tooltip-hero {
        visibility: visible;
        opacity: 1;
    }
    </style>
""", unsafe_allow_html=True)

# --- THUẬT TOÁN ĐỊNH DẠNG SỐ VIỆT HÓA CHUẨN TRÁNH TRÙNG LẶP DO LỖI GÕ KÝ TỰ ---
def clean_vietnamese_number(val):
    if pd.isna(val):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace('\xa0', '').replace(' ', '')
    if not s:
        return 0.0
    
    # Chuẩn hóa tiền tệ và ký hiệu đặc biệt
    s = s.replace('₫', '').replace('đ', '').replace('VND', '').replace('vnd', '').strip()
    
    # Định dạng "380.952.000,50" -> chuyển dấu chấm thành rỗng, dấu phẩy thành dấu chấm thập phân
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    elif '.' in s:
        parts = s.split('.')
        # Nếu có nhiều hơn 1 dấu chấm hoặc phần cuối cùng có độ dài đúng bằng 3 chữ số -> Dấu chấm là phân tách hàng nghìn
        if len(parts) > 2 or (len(parts) == 2 and len(parts[1]) == 3):
            s = "".join(parts)
            
    try:
        return float(s)
    except:
        return 0.0

# --- THUẬT TOÁN CHUẨN HÓA UNICODE TIẾNG VIỆT & LỌC SAI LỖI CHÍNH TẢ TỪNG DÒNG EXCEL (Mục 1 & 2) ---
def clean_customer_name(text):
    if pd.isna(text):
        return ""
    # Chuẩn hóa khoảng cách và định dạng NFC chuẩn của hệ thống tiếng Việt
    s = " ".join(str(text).split())
    s_norm = unicodedata.normalize('NFC', s)
    
    # Khai báo bộ từ điển ánh xạ chuẩn hóa chính tả đặc hiệu cho dữ liệu gốc (Hợp nhất Hoàng Oánh & Medilab)
    typo_dict = {
        "hoaàng oánh": "Hoàng Oánh",
        "hoàng oánh": "Hoàng Oánh",
        "hoàng oanh": "Hoàng Oánh",
        "medilâb": "Medilab",
        "medilab": "Medilab",
        "tbyt linhken": "TBYT Linhken"
    }
    
    s_lower = s_norm.lower()
    if s_lower in typo_dict:
        return typo_dict[s_lower]
        
    return s_norm

# ==========================================
# 2. HỆ THỐNG XỬ LÝ DỮ LIỆU CHUẨN HÓA
# ==========================================
@st.cache_data(ttl=600)
def load_data():
    sheet_id = st.secrets["spreadsheet_id"]
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    
    df_th_raw = pd.read_excel(url, sheet_name='Tổng hợp', header=1)
    
    # 1. Tìm vị trí cột HSD
    hsd_col_idx = None
    for idx, col in enumerate(df_th_raw.columns):
        if any(keyword in str(col).lower() for keyword in ['hsd', 'hạn', 'date', 'quá hạn']):
            hsd_col_idx = idx
            break
            
    if hsd_col_idx is None and df_th_raw.shape[1] > 17:
        hsd_col_idx = 17

    # 2. Tìm vị trí cột Đơn vị tính (ĐVT) tự động từ cấu trúc File
    dvt_col_idx = None
    for idx, col in enumerate(df_th_raw.columns):
        if any(keyword in str(col).lower() for keyword in ['đvt', 'đơn vị tính', 'dvt', 'unit', 'đơn vị']):
            dvt_col_idx = idx
            break

    # 3. Tạo ánh xạ cột an toàn để tránh ghi đè chỉ số cột
    col_mapping = {
        1: 'Nganh_Hang',
        2: 'Chung_Loai',
        3: 'Hang',
        4: 'SKU',
        10: 'Xuat_Ban_SL',
        14: 'Ton_Kho_SL',
        15: 'Ton_Kho_Value'
    }
    
    if hsd_col_idx is not None and hsd_col_idx not in col_mapping:
        col_mapping[hsd_col_idx] = 'Het_HSD_Value'
        
    if dvt_col_idx is not None and dvt_col_idx not in col_mapping:
        col_mapping[dvt_col_idx] = 'DVT'

    # Lọc và đổi tên cột
    sorted_keys = sorted(list(col_mapping.keys()))
    df_th = df_th_raw.iloc[:, sorted_keys].copy()
    df_th.columns = [col_mapping[k] for k in sorted_keys]
    
    # Gán các giá trị mặc định phòng trường hợp cột thiếu dữ liệu
    if 'Het_HSD_Value' not in df_th.columns:
        df_th['Het_HSD_Value'] = 0.0
    if 'DVT' not in df_th.columns:
        df_th['DVT'] = 'Cái'
        
    df_th = df_th.dropna(subset=['SKU'])
    
    # Áp dụng chuẩn hóa Unicode NFC chống trùng lặp nhóm
    df_th['SKU'] = df_th['SKU'].apply(clean_customer_name)
    
    for col in ['Nganh_Hang', 'Chung_Loai', 'Hang']:
        if col in df_th.columns: 
            df_th[col] = df_th[col].fillna('Khác').astype(str).str.strip().apply(clean_customer_name)
            
    df_th['DVT'] = df_th['DVT'].fillna('Cái').astype(str).str.strip().apply(clean_customer_name)
    
    for col in ['Xuat_Ban_SL', 'Ton_Kho_SL', 'Ton_Kho_Value', 'Het_HSD_Value']:
        if col in df_th.columns: 
            df_th[col] = pd.to_numeric(df_th[col], errors='coerce').fillna(0)

    # --- ĐỌC VÀ LỌC DỮ LIỆU SHEET XUẤT BÁN ---
    df_xb_raw = pd.read_excel(url, sheet_name='Xuất bán', header=1)
    
    # Chuẩn hóa cột Mã SKU (Cột B - Index 1)
    cols = df_xb_raw.columns.tolist()
    cols[1] = 'SKU'
    df_xb_raw.columns = cols
    
    # Tìm chính xác cột Khách hàng (Cột F - Index 5)
    kh_cols = [c for c in df_xb_raw.columns if 'Khách hàng' in str(c)]
    kh_col_name = kh_cols[0] if kh_cols else df_xb_raw.columns[5]
    
    # Tìm chính xác cột ĐVT (Cột C - Index 2)
    dvt_cols = [c for c in df_xb_raw.columns if 'Đvt' in str(c) or 'ĐVT' in str(c)]
    dvt_col_name = dvt_cols[0] if dvt_cols else df_xb_raw.columns[2]
    
    # Tìm chính xác cột Số lượng (Cột D - Index 3)
    qty_cols = [c for c in df_xb_raw.columns if 'Số lượng' in str(c)]
    qty_col_name = qty_cols[0] if qty_cols else df_xb_raw.columns[3]
    
    # Tìm chính xác cột Doanh thu (Cột E - Index 4)
    val_cols = [c for c in df_xb_raw.columns if 'Doanh thu' in str(c)]
    val_col_name = val_cols[0] if val_cols else df_xb_raw.columns[4]
    
    # Tiến hành chuẩn hóa lọc và làm sạch dữ liệu xuất bán theo đúng cột gốc
    df_xb_clean = df_xb_raw.copy()
    df_xb_clean.rename(columns={
        kh_col_name: 'Khach_Hang',
        dvt_col_name: 'DVT_Xuat',
        qty_col_name: 'Quantity',
        val_col_name: 'Value'
    }, inplace=True)
    
    # Áp dụng chuẩn hóa Unicode và lọc sạch lỗi chính tả cho cột Khách hàng (Mục 2)
    df_xb_clean['Khach_Hang'] = df_xb_clean['Khach_Hang'].apply(clean_customer_name)
    df_xb_clean['SKU'] = df_xb_clean['SKU'].apply(clean_customer_name)
    
    # Loại bỏ các dòng ghi chú tổng hợp trống
    df_xb_clean = df_xb_clean[df_xb_clean['Khach_Hang'] != '']
    
    # SỬA LỖI MẤT DỮ LIỆU: Chuẩn hóa gộp nhóm không phân biệt hoa thường để tránh Hoàng Oánh bị chia tách
    df_xb_clean['Khach_Hang_Lower'] = df_xb_clean['Khach_Hang'].str.lower()
    proper_case_map = df_xb_clean.groupby('Khach_Hang_Lower')['Khach_Hang'].first().to_dict()
    df_xb_clean['Khach_Hang'] = df_xb_clean['Khach_Hang_Lower'].map(proper_case_map)
    
    # Giải quyết triệt để lỗi chuyển đổi chuỗi số từ Excel tiếng Việt
    df_xb_clean['Value'] = df_xb_clean['Value'].apply(clean_vietnamese_number)
    df_xb_clean['Quantity'] = df_xb_clean['Quantity'].apply(clean_vietnamese_number)

    # --- SỬA LỖI LỌC DỮ LIỆU THÁNG GIAO DỊCH (Cho phép đọc chuỗi text "Tháng xx/yyyy") ---
    thang_cols = [c for c in df_xb_raw.columns if 'Tháng' in str(c)]
    if thang_cols:
        thang_col_name = thang_cols[0]

        def parse_thang_to_date(val):
            if pd.isna(val):
                return pd.NaT
            if isinstance(val, (datetime.datetime, datetime.date, pd.Timestamp)):
                return pd.to_datetime(val)
                
            val_str = str(val).strip().lower()
            # Loại bỏ chữ "tháng" hoặc khoảng trắng thừa trong chuỗi text
            val_str = val_str.replace('tháng', '').replace('thang', '').strip()
            try:
                if '/' in val_str:
                    parts = val_str.split('/')
                    if len(parts) == 2:
                        m = int(parts[0].strip())
                        y = int(parts[1].strip())
                        return pd.to_datetime(f"01/{m:02d}/{y}", format="%d/%m/%Y")
                if len(val_str) == 6 and val_str.isdigit():
                    return pd.to_datetime(val_str, format="%Y%m")
                return pd.to_datetime(val_str, errors='coerce')
            except:
                return pd.NaT

        df_xb_clean['Date_Filter'] = df_xb_clean[thang_col_name].apply(parse_thang_to_date)
        df_xb_clean = df_xb_clean[df_xb_clean['Date_Filter'] >= pd.to_datetime('2026-01-01')]
    
    customer_mapping = df_xb_clean[['SKU', 'Khach_Hang']].copy()

    # Số lượng khách hàng active theo từng SKU
    active_kh = df_xb_clean.groupby('SKU')['Khach_Hang'].nunique().reset_index()
    active_kh.rename(columns={'Khach_Hang': 'Khach_Hang_Active'}, inplace=True)
    
    df = pd.merge(df_th, active_kh, on='SKU', how='left').fillna(0)
    return df, customer_mapping, df_xb_clean

# --- THUẬT TOÁN ĐỊNH DẠNG FILE PO EXCEL BẰNG OPENPYXL (MẶC ĐỊNH SẴN CÓ) ---
def to_excel(df_to_export):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_to_export.to_excel(writer, index=False, sheet_name='SGM_PO_Export')
        workbook = writer.book
        worksheet = writer.sheets['SGM_PO_Export']
        
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Định dạng Style màu xanh thương hiệu SGM
        header_fill = PatternFill(start_color="388E3C", end_color="388E3C", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell_font = Font(name="Segoe UI", size=10)
        
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )
        
        # Thiết kế Header
        for col_idx in range(1, len(df_to_export.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        # Thiết kế nội dung bảng dữ liệu
        for row_idx in range(2, len(df_to_export) + 2):
            for col_idx in range(1, len(df_to_export.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = cell_font
                cell.border = thin_border
                
                val = cell.value
                if isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                    # Định dạng phân tách hàng nghìn cho số lượng và giá trị (Ngoại trừ Khách Hàng Active)
                    if "Active" not in df_to_export.columns[col_idx-1]:
                        cell.number_format = '#,##0'
                else:
                    cell.alignment = Alignment(horizontal="left")
                    
        # Tự động điều chỉnh độ rộng cột
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)
            
    return output.getvalue()

# ==========================================
# 3. GIAO DIỆN CHÍNH & THUẬT TOÁN ĐIỀU PHỐI
# ==========================================
try:
    df_full, customer_mapping, df_xb_clean = load_data()
    today = datetime.date.today()
    
    # --- SIDEBAR MENU ---
    st.sidebar.image("https://raw.githubusercontent.com/vic-techSGM/sgm-supplychain/main/logo.png", use_container_width=True)
    st.sidebar.markdown("<div class='copyright-text'>Built & Developed by Vic Fan.<br>All rights reserved. Version 01.26.05</div>", unsafe_allow_html=True)
    
    st.sidebar.markdown("<h4>⚙️ THAM SỐ DỰ TRÙ</h4>", unsafe_allow_html=True)
    doi_target = st.sidebar.slider("Ngày tồn kho an toàn (DOI)", 15, 120, 45)
    st.sidebar.caption("💡15 (Nguy cơ) - 45 (Chuẩn) - 90 (Đọng vốn)")
    lead_time = st.sidebar.slider("Lead Time (Ngày nhập)", 10, 90, 30)
    customer_growth = st.sidebar.slider("Kỳ vọng tăng trưởng Khách hàng theo Quý (%)", 0, 100, 15)
    
    st.sidebar.markdown("<h4>📂 BỘ LỌC</h4>", unsafe_allow_html=True)
    list_nganh = df_full['Nganh_Hang'].unique().tolist()
    selected_nganh = st.sidebar.multiselect("Ngành hàng", list_nganh, default=list_nganh)
    
    df_f1 = df_full[df_full['Nganh_Hang'].isin(selected_nganh)]
    list_chungloai = df_f1['Chung_Loai'].unique().tolist() if 'Chung_Loai' in df_f1.columns else []
    selected_chungloai = st.sidebar.multiselect("Chủng loại", list_chungloai, default=list_chungloai) if list_chungloai else []
    
    df_f2 = df_f1[df_f1['Chung_Loai'].isin(selected_chungloai)] if selected_chungloai else df_f1
    list_hang = df_f2['Hang'].unique().tolist()
    selected_hang = st.sidebar.multiselect("Hãng", list_hang, default=list_hang)

    df = df_f2[df_f2['Hang'].isin(selected_hang)].copy()
    
    # --- THUẬT TOÁN TÍNH TOÁN ROP VÀ S2S ---
    growth_factor = 1 + (customer_growth / 100)
    df['Daily_Sales'] = (df['Xuat_Ban_SL'] / 150) * growth_factor 
    df['S2S_Months'] = df['Ton_Kho_SL'] / ((df['Daily_Sales'] * 30) + 0.0001)
    
    df['Du_Tru_Thang'] = df['Daily_Sales'] * 30
    df['Du_Tru_Quy'] = df['Daily_Sales'] * 90
    
    df['ROP_Qty'] = (lead_time * df['Daily_Sales']) + (doi_target * df['Daily_Sales'])
    df['De_Xuat_Mua'] = (df['ROP_Qty'] - df['Ton_Kho_SL']).apply(lambda x: max(int(x), 0))

    def calculate_reorder_date(row):
        if row['Daily_Sales'] <= 0: return "Chưa phát sinh"
        days_to_rop = (row['Ton_Kho_SL'] - row['ROP_Qty']) / row['Daily_Sales']
        if days_to_rop <= 0: return today.strftime("%d/%m/%Y") 
        future_date = today + datetime.timedelta(days=int(days_to_rop))
        return future_date.strftime("%d/%m/%Y")
        
    df['Ngay_Dat_Hang_Du_Kien'] = df.apply(calculate_reorder_date, axis=1)

    def get_status(row):
        if row['Ton_Kho_SL'] < (lead_time * row['Daily_Sales']): return "🔴 ĐỨT HÀNG"
        if row['Ton_Kho_SL'] < row['ROP_Qty']: return "🟡 CẦN NHẬP"
        return "🟢 AN TOÀN"
    
    def get_s2s_alert(row):
        if row['S2S_Months'] > 6: return "⚠️ Chậm luân chuyển"
        if row['S2S_Months'] < 1: return "🔥 Rủi ro thiếu hàng"
        return "✅ Hợp lý"

    df['Trang_Thai'] = df.apply(get_status, axis=1)
    df['Canh_Bao_S2S'] = df.apply(get_s2s_alert, axis=1)

    total_daily_sales = df['Daily_Sales'].sum()
    avg_s2s_global = df['Ton_Kho_SL'].sum() / ((total_daily_sales * 30) + 0.0001) if total_daily_sales > 0 else 0

    # --- GIẢI QUYẾT TRÙNG LẶP: Tính toán lượng khách hàng active duy nhất theo bộ lọc đang chạy ---
    active_skus = df['SKU'].unique()
    filtered_customers = customer_mapping[customer_mapping['SKU'].isin(active_skus)]
    total_active_customers_clean = filtered_customers['Khach_Hang'].nunique()

    # --- HERO SECTION TỔNG QUAN ---
    st.markdown("<h2 style='font-weight: 900; margin-bottom: 5px; color: #0f172a;'>HỆ THỐNG QUẢN TRỊ KHO & GIÁM SÁT CUNG ỨNG</h2>", unsafe_allow_html=True)
    
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("TỔNG VỐN TỒN KHO", f"{df['Ton_Kho_Value'].sum():,.0f} ₫")
    m2.metric("SỐ LƯỢNG SKU", f"{len(df):,}")
    
    # Thiết kế lại m3 sử dụng thẻ HTML có tích hợp Tooltip (Đuy tooltip lên trên đầu)
    with m3:
        st.markdown(f"""
        <div class="custom-hero-card">
            <div class="hero-label">S2S BÌNH QUÂN</div>
            <div class="hero-value">{avg_s2s_global:.1f} Tháng</div>
            <div class="tooltip-hero">
                <strong style="color: #fb923c; font-size: 13px;">Chỉ số S2S (Stock-to-Sales) Bình quân:</strong><br><br>
                • <b>Định nghĩa S2S:</b> Biểu thị số tháng cần thiết để phân phối hết lượng hàng hóa đang lưu kho thực tế dựa trên tốc độ tiêu thụ hiện tại.<br>
                • <b>Thuật toán:</b> Bằng Tổng sản lượng tồn kho của toàn bộ hệ thống chia cho Tổng doanh số bán hàng trung bình mỗi tháng.<br>
                • <b>Ý nghĩa điều phối:</b> Hỗ trợ nhà quản trị nhận định nhanh thời gian quay vòng của toàn bộ nguồn vốn đang lưu chuyển trong kho.
            </div>
        </div>
        """, unsafe_allow_html=True)
        
    m4.metric("KHÁCH HÀNG CÓ GIAO DỊCH", f"{int(total_active_customers_clean):,}")

    st.markdown("<br>", unsafe_allow_html=True)
    
    # --- ĐIỀU HƯỚNG TABS ---
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📊 TỔNG QUAN & DỰ TRÙ", 
        "👥 KHÁCH HÀNG THEO SKU", 
        "🚩 RỦI RO HẠN DÙNG",
        "📈 PHÂN LOẠI SKU THEO DOANH THU", 
        "💡 TRA CỨU CHI TIẾT SKU"
    ])

    # --- TAB 1: TỔNG QUAN & ĐẶT HÀNG ---
    with tab1:
        st.markdown("<h4 style='font-weight: 800; margin-top: 10px;'>📊 Tổng Quan Cơ Cấu Vốn</h4>", unsafe_allow_html=True)
        col_p1, col_p2 = st.columns(2)
        with col_p1:
            fig_pie_nganh = px.pie(
                df, values='Ton_Kho_Value', names='Nganh_Hang', hole=0.4, 
                title="Phân bổ vốn theo Ngành Hàng", 
                color_discrete_sequence=px.colors.qualitative.Pastel
            )
            # Tối ưu hóa Tooltip Việt hóa cho biểu đồ ngành
            fig_pie_nganh.update_traces(
                textposition='inside', 
                textinfo='percent',
                hovertemplate="<b>Ngành hàng:</b> %{label}<br><b>Giá trị tồn kho:</b> %{value:,.0f} ₫<br><b>Tỷ trọng:</b> %{percent}<extra></extra>"
            )
            fig_pie_nganh.update_layout(
                showlegend=True,
                legend=dict(
                    orientation="h",
                    yanchor="top",
                    y=-0.1,
                    xanchor="center",
                    x=0.5,
                    font=dict(size=9)
                ),
                paper_bgcolor='rgba(0,0,0,0)', 
                plot_bgcolor='rgba(0,0,0,0)',
                margin=dict(t=50, b=50, l=10, r=10),
                font=dict(family="Montserrat", color="#1e293b")
            )
            st.plotly_chart(fig_pie_nganh, use_container_width=True)
            
        with col_p2:
            fig_pie_hang = px.pie(
                df, values='Ton_Kho_Value', names='Hang', hole=0.4, 
                title="Tỷ trọng vốn theo Hãng", 
                color_discrete_sequence=px.colors.qualitative.Set2
            )
            # Tối ưu hóa Tooltip Việt hóa cho biểu đồ hãng, hỗ trợ Scale responsive
            fig_pie_hang.update_traces(
                textposition='inside', 
                textinfo='percent',
                hovertemplate="<b>Hãng sản xuất:</b> %{label}<br><b>Giá trị tồn kho:</b> %{value:,.0f} ₫<br><b>Tỷ trọng:</b> %{percent}<extra></extra>"
            )
            fig_pie_hang.update_layout(
                showlegend=True,
                legend=dict(
                    orientation="h",
                    yanchor="top",
                    y=-0.1,  
                    xanchor="center",
                    x=0.5,
                    font=dict(size=9) 
                ),
                paper_bgcolor='rgba(0,0,0,0)', 
                plot_bgcolor='rgba(0,0,0,0)',
                margin=dict(t=50, b=50, l=10, r=10), 
                font=dict(family="Montserrat", color="#1e293b")
            )
            st.plotly_chart(fig_pie_hang, use_container_width=True)
            
        st.markdown("<p style='text-align: center; font-size: 13px; color: #64748b;'><i>*Trỏ chuột vào biểu đồ để xem chi tiết tên hãng/ngành và giá trị vốn cụ thể</i></p>", unsafe_allow_html=True)

        st.markdown("<h4 style='font-weight: 800; margin-top: 30px;'>🛒 Thống kê dự trù</h4>", unsafe_allow_html=True)
        
        # Bổ sung trường dữ liệu ĐVT vào bảng Thống kê dự trù
        display_df = df[['SKU', 'Hang', 'DVT', 'Ton_Kho_SL', 'Khach_Hang_Active', 'Du_Tru_Thang', 'Du_Tru_Quy', 'Ngay_Dat_Hang_Du_Kien', 'De_Xuat_Mua', 'Trang_Thai', 'Canh_Bao_S2S']].copy()
        display_df.columns = [
            'Mã SKU', 'Hãng', 'ĐVT', 'Số Lượng Tồn Kho', 'Khách Hàng Active', 
            'Dự Trù Trong Tháng', 'Dự Trù 3 Tháng (Quý)', 'Ngày Đặt Hàng (ROP)', 
            'Số Lượng Cần Mua', 'Trạng Thái', 'Cảnh Báo S2S'
        ]
        
        # Định nghĩa kiểu dữ liệu cột Khách Hàng Active là kiểu Số Nguyên thực tế
        display_df['Khách Hàng Active'] = display_df['Khách Hàng Active'].astype(int)
        
        # SỬA LỖI HIỂN THỊ TOOLTIP HOÀN TOÀN BẰNG CẤU HÌNH COLUMN_CONFIG
        # Giữ đúng kiểu NumberColumn nhưng sử dụng format="%d" để không áp dụng phân tách hàng nghìn
        # Bổ sung tooltip tiếng Việt chi tiết cho từng header cột tương ứng, bao gồm ĐVT
        st.dataframe(
            display_df, 
            use_container_width=True, 
            height=350,
            hide_index=True,
            column_config={
                "Mã SKU": st.column_config.TextColumn("Mã SKU", help="Mã định danh duy nhất của sản phẩm"),
                "Hãng": st.column_config.TextColumn("Hãng", help="Thương hiệu/Nhà sản xuất sản phẩm"),
                "ĐVT": st.column_config.TextColumn("ĐVT", help="Đơn vị tính của sản phẩm (Chai, Hộp, Bộ...)"),
                "Số Lượng Tồn Kho": st.column_config.NumberColumn("Số Lượng Tồn Kho", format="%d", help="Số lượng sản phẩm hiện có trong kho thực tế"),
                "Khách Hàng Active": st.column_config.NumberColumn("Khách Hàng Active", format="%d", help="Số lượng khách hàng duy nhất đã mua mặt hàng này tính từ năm 2026"), 
                "Dự Trù Trong Tháng": st.column_config.NumberColumn("Dự Trù Trong Tháng", format="%d", help="Nhu cầu sản lượng dự kiến tiêu thụ trong 30 ngày tiếp theo"),
                "Dự Trù 3 Tháng (Quý)": st.column_config.NumberColumn("Dự Trù 3 Tháng (Quý)", format="%d", help="Nhu cầu sản lượng dự kiến tiêu thụ trong 90 ngày tiếp theo (quý)"),
                "Ngày Đặt Hàng (ROP)": st.column_config.TextColumn("Ngày Đặt Hàng (ROP)", help="Thời điểm tối ưu cần thực hiện đặt hàng PO mới"),
                "Số Lượng Cần Mua": st.column_config.NumberColumn("Số Lượng Cần Mua", format="%d", help="Số lượng tối ưu cần lên đơn mua bổ sung"),
                "Trạng Thái": st.column_config.TextColumn("Trạng Thái", help="Đánh giá mức độ an toàn của kho hàng"),
                "Cảnh Báo S2S": st.column_config.TextColumn("Cảnh Báo S2S", help="Vòng quay luân chuyển hàng tồn kho")
            }
        )
        
        st.markdown("""
        <div class="smart-card-info">
            <b style="color:#1d4ed8;">ℹ️ CƠ CHẾ DỰ BÁO NGÀY ĐẶT HÀNG (ROP DATE):</b><br>
            Hệ thống tự động ước lượng <b>Thời điểm chính xác cần xuống đơn PO (Ngày/Tháng/Năm)</b>. Ngày dự kiến được hệ thống tính toán dựa vào 3 tham số thiết lập: <b>Tốc độ tăng trưởng KH</b>, thời gian <b>Lead Time</b> chở hàng về, và số ngày tồn trữ phòng hờ <b>DOI</b>.
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("<h4 style='font-weight: 800; margin-top: 30px;'>📥 Xuất File Đơn Đặt Hàng (PO)</h4>", unsafe_allow_html=True)
        po_df = display_df[display_df['Số Lượng Cần Mua'] > 0].copy()
        
        if not po_df.empty:
            export_cols = st.multiselect(
                "Chọn các cột dữ liệu muốn trích xuất:", 
                options=po_df.columns.tolist(), 
                default=po_df.columns.tolist()
            )
            
            # Xuất PO bằng openpyxl (mặc định trong Streamlit), format chuẩn ready-to-use
            excel_data = to_excel(po_df[export_cols])
            
            st.download_button(
                label="📥 TẢI XUỐNG FILE PO CHUẨN EXCEL (.XLSX)", 
                data=excel_data, 
                file_name=f'SGM_PO_{today.strftime("%Y%m%d")}.xlsx', 
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            st.success("Không có mặt hàng nào cần đặt mua trong danh mục lọc hiện tại.")

    # --- TAB 2: KHÁCH HÀNG THEO SKU ---
    with tab2:
        st.markdown("<h3 style='font-weight: 800;'>🔍 Phân tích Hành vi Khách hàng theo SKU</h3>", unsafe_allow_html=True)

        # 1. Thuật toán tính toán tự động TOP 5 Khách hàng có doanh thu cao nhất cho đến thời điểm up-to-date
        top_5_cust = df_xb_clean.groupby('Khach_Hang')['Value'].sum().reset_index()
        top_5_cust = top_5_cust.sort_values(by='Value', ascending=False).head(5)
        
        st.markdown("<h4 style='font-weight: 800; margin-top: 15px; margin-bottom: 12px; color: #1e293b;'>👑 TOP 5 KHÁCH HÀNG ĐANG CÓ DOANH THU CAO NHẤT</h4>", unsafe_allow_html=True)
        
        # 2. Xây dựng Layout hàng ngang gồm 5 cột hiển thị tên và doanh thu tổng của từng Khách hàng
        cols_top = st.columns(5)
        for idx, row in enumerate(top_5_cust.iterrows()):
            with cols_top[idx]:
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #f8fafc, #eff6ff); border: 1px solid #bfdbfe; border-top: 4px solid #3b82f6; padding: 12px 15px; border-radius: 10px; min-height: 85px; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.04);">
                    <div style="color: #1e3a8a; font-size: 11px; font-weight: 800; text-transform: uppercase;">Hạng {idx+1}</div>
                    <div style="color: #334155; font-size: 13px; font-weight: 700; margin-top: 4px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap;" title="{row[1]['Khach_Hang']}">{row[1]['Khach_Hang']}</div>
                    <div style="color: #2563eb; font-size: 14px; font-weight: 800; margin-top: 2px;">{row[1]['Value']:,.0f} ₫</div>
                </div>
                """, unsafe_allow_html=True)
        
        st.markdown("<br>", unsafe_allow_html=True)

        # 3. Thuật toán tính toán nhóm 10 khách hàng có doanh thu thấp nhất (Đã cập nhật chuẩn hóa dồn nhóm - Hoàng Oánh hiển thị đúng mốc 47 triệu)
        all_cust_rev = df_xb_clean.groupby('Khach_Hang')['Value'].sum().reset_index()
        bottom_10_cust = all_cust_rev[all_cust_rev['Value'] > 0].sort_values(by='Value', ascending=True).head(10)
        
        if not bottom_10_cust.empty:
            fig_bottom = px.bar(
                bottom_10_cust,
                x='Value',
                y='Khach_Hang',
                orientation='h',
                text='Value',
                title="Nhóm 10 Khách hàng có doanh thu thấp nhất (Cần lưu ý thúc đẩy)",
                labels={'Value': 'Doanh thu lũy kế (₫)', 'Khach_Hang': 'Khách hàng'},
                color='Value',
                color_continuous_scale=px.colors.sequential.Reds
            )
            fig_bottom.update_traces(
                texttemplate='%{text:,.0f} ₫',
                textposition='outside',
                hovertemplate="<b>Khách hàng:</b> %{y}<br><b>Doanh thu tích lũy:</b> %{x:,.0f} ₫<extra></extra>"
            )
            fig_bottom.update_layout(
                coloraxis_showscale=False,
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)',
                font=dict(family="Montserrat", color="#1e293b"),
                yaxis=dict(autorange="reversed")  # Sắp xếp từ thấp nhất ở phía trên cùng
            )
            st.plotly_chart(fig_bottom, use_container_width=True)
            st.markdown("<br>", unsafe_allow_html=True)

        # 4. Đẩy hộp tìm kiếm đa lựa chọn xuống dưới
        list_kh = sorted(customer_mapping['Khach_Hang'].unique().tolist())
        selected_kh = st.multiselect("Gõ tên hoặc chọn để tìm kiếm Khách Hàng (có thể chọn nhiều để so sánh):", list_kh)
        
        # --- THUẬT TOÁN KHI KHÔNG CHỌN KHÁCH HÀNG NÀO ---
        if not selected_kh:
            st.info("Vui lòng gõ tên hoặc chọn ít nhất một Khách Hàng từ hộp tìm kiếm phía trên để hiển thị phân tích dữ liệu.")

        # --- THUẬT TOÁN HIỂN THỊ CHI TIẾT KHI CHỌN DUY NHẤT 1 KHÁCH HÀNG (Chuẩn hóa thông tin hiển thị 4 thuộc tính - Mục 1 & 2) ---
        elif len(selected_kh) == 1:
            cust = selected_kh[0]
            cust_tx = df_xb_clean[df_xb_clean['Khach_Hang'] == cust].copy()
            
            if not cust_tx.empty:
                # 1. Định dạng lịch sử tháng đặt hàng (mỗi tháng 1 dòng, không bold)
                cust_tx['Thang_Str'] = cust_tx['Date_Filter'].dt.strftime('%m/%Y')
                ordered_months = cust_tx.sort_values('Date_Filter')['Thang_Str'].dropna().unique().tolist()
                months_history_html = "<br>".join(ordered_months) if ordered_months else "Chưa có lịch sử"
                
                # 2. Tính giá trị đơn hàng bình quân và xây dựng Tooltip hiển thị doanh thu theo từng tháng
                avg_order_val = cust_tx.groupby('Date_Filter')['Value'].sum().mean()
                if pd.isna(avg_order_val):
                    avg_order_val = 0.0
                
                # Tính tổng chi tiêu theo tháng để hiển thị trên Tooltip của Card Đơn hàng trung bình
                monthly_totals = cust_tx.groupby('Thang_Str')['Value'].sum().reset_index()
                monthly_totals['Date_Sort'] = pd.to_datetime(monthly_totals['Thang_Str'], format='%m/%Y')
                monthly_totals = monthly_totals.sort_values('Date_Sort')
                
                monthly_totals_html = ""
                for _, row in monthly_totals.iterrows():
                    monthly_totals_html += f"• Tháng {row['Thang_Str']}: {row['Value']:,.0f} ₫<br>"
                
                if not monthly_totals_html:
                    monthly_totals_html = "Chưa ghi nhận dữ liệu doanh thu"

                # Ánh xạ lấy Hãng và ĐVT tương ứng từ bảng Tổng hợp
                sku_brand_map = df.set_index('SKU')['Hang'].to_dict()
                sku_dvt_map = df.set_index('SKU')['DVT'].to_dict()

                # 3. Thuật toán đo lường chuỗi đặt hàng liên tục (Streak) không đứt quãng theo tháng cho Top 5 SKU
                def get_max_streak(dates_series):
                    if dates_series.empty:
                        return 0
                    valid_dates = [d for d in dates_series if pd.notna(d)]
                    months_indices = []
                    for d in valid_dates:
                        months_indices.append(d.year * 12 + d.month)
                    indices = sorted(list(set(months_indices)))
                    if not indices:
                        return 0
                    max_streak = 1
                    current_streak = 1
                    for i in range(1, len(indices)):
                        if indices[i] == indices[i-1] + 1:
                            current_streak += 1
                        else:
                            max_streak = max(max_streak, current_streak)
                            current_streak = 1
                    return max(max_streak, current_streak)

                sku_streaks = {}
                sku_months = cust_tx.groupby('SKU')['Date_Filter'].apply(list).to_dict()
                for sku, dates in sku_months.items():
                    sku_streaks[sku] = get_max_streak(pd.Series(dates))
                
                sku_stats = cust_tx.groupby('SKU').agg(
                    total_qty=('Quantity', 'sum')
                ).reset_index()
                sku_stats['streak'] = sku_stats['SKU'].map(sku_streaks).fillna(0)
                
                sku_stats = sku_stats.sort_values(by=['streak', 'total_qty'], ascending=[False, False])
                top_5_skus = sku_stats.head(5)['SKU'].tolist()
                top_5_html = "<br>".join(top_5_skus) if top_5_skus else "Chưa ghi nhận"

                # Hiển thị khối 3 thẻ thông số (Customer Insight Cards)
                col_c1, col_c2, col_c3 = st.columns(3)
                with col_c1:
                    st.markdown(f"""
                    <div style="background: linear-gradient(145deg, #f8fafc, #f1f5f9); border: 1px solid #cbd5e1; border-left: 5px solid #3b82f6; padding: 20px; border-radius: 12px; min-height: 120px;">
                        <span style="color: #475569; font-size: 11px; text-transform: uppercase; font-weight: 800; letter-spacing: 0.5px;">📅 Lịch sử tháng đặt hàng</span>
                        <p style="margin: 8px 0 0 0; color: #1e293b; font-size: 14px; font-weight: 500; line-height: 1.5;">{months_history_html}</p>
                    </div>
                    """, unsafe_allow_html=True)
                with col_c2:
                    st.markdown(f"""
                    <div style="background: linear-gradient(145deg, #f8fafc, #f1f5f9); border: 1px solid #cbd5e1; border-left: 5px solid #10b981; padding: 20px; border-radius: 12px; min-height: 120px;">
                        <span style="color: #475569; font-size: 11px; text-transform: uppercase; font-weight: 800; letter-spacing: 0.5px;">🔥 Top 5 SKU mua liên tục</span>
                        <p style="margin: 8px 0 0 0; color: #1e293b; font-size: 14px; font-weight: 500; line-height: 1.5;">{top_5_html}</p>
                    </div>
                    """, unsafe_allow_html=True)
                with col_c3:
                    st.markdown(f"""
                    <div class="custom-metric-card-aov">
                        <span style="color: #475569; font-size: 11px; text-transform: uppercase; font-weight: 800; letter-spacing: 0.5px;">💰 Giá trị đơn trung bình</span>
                        <p style="margin: 8px 0 0 0; color: #1e293b; font-size: 22px; font-weight: 900; line-height: 1.2;">{avg_order_val:,.0f} ₫</p>
                        <div class="tooltip-text-aov">
                            <strong style="color: #f59e0b; font-size: 13px;">Doanh thu đặt hàng từng tháng:</strong><br><br>
                            {monthly_totals_html}
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                st.markdown("<div style='margin-bottom: 20px;'></div>", unsafe_allow_html=True)

                # --- PHẦN KHAI BÁO PHÂN TÍCH GIAO DỊCH CHI TIẾT THEO YÊU CẦU ĐỒNG BỘ 4 THUỘC TÍNH (Mục 1) ---
                total_cust_spend = cust_tx['Value'].sum()
                cust_skus_count = cust_tx['SKU'].nunique()
                
                # Tìm mốc giao dịch có doanh thu lớn nhất của khách hàng
                largest_idx = cust_tx['Value'].idxmax()
                largest_row = cust_tx.loc[largest_idx]
                largest_val = largest_row['Value']
                largest_sku = largest_row['SKU'] # SKU chứa mô tả tên đầy đủ của mặt hàng
                largest_brand = sku_brand_map.get(largest_sku, "Khác")
                largest_dvt = sku_dvt_map.get(largest_sku, largest_row['DVT_Xuat'])
                largest_month = largest_row['Date_Filter'].strftime('%m/%Y')
                
                # Tìm top 4 giao dịch lớn tiếp theo xếp hạng giảm dần
                other_txs = cust_tx.drop(index=largest_idx).sort_values(by='Value', ascending=False).head(4)
                other_tx_lines = []
                for _, row in other_txs.iterrows():
                    r_sku = row['SKU']
                    r_brand = sku_brand_map.get(r_sku, "Khác")
                    r_dvt = sku_dvt_map.get(r_sku, row['DVT_Xuat'])
                    r_val = row['Value']
                    r_month = row['Date_Filter'].strftime('%m/%Y')
                    # Đưa về định dạng chuẩn hoá 4 thuộc tính: Tên SKU — Hãng sản xuất — Đơn vị tính — Giá trị giao dịch (Tháng đặt hàng) (Mục 1)
                    other_tx_lines.append(f"<b>{r_sku}</b> — <b>{r_brand}</b> — <b>{r_dvt}</b> — <b>{r_val:,.0f} ₫</b> (Tháng {r_month})")
                other_tx_html = "<br>".join(other_tx_lines) if other_tx_lines else "Không có giao dịch tiêu biểu khác"

                # Hiển thị cấu trúc mô tả chi tiết giao dịch theo chuẩn hoá cấu trúc phẳng (Mục 1)
                st.markdown(f"""
                <div class="smart-card-success">
                    <b style="color:#15803d; font-size:16px;">📊 ĐÁNH GIÁ KHÁCH HÀNG:</b><br><br>
                    • Khách hàng <b>{cust}</b> đem lại doanh thu lũy kế <b>{total_cust_spend:,.0f} ₫</b> trên tổng số <b>{cust_skus_count} mã SKU</b>.<br><br>
                    • <b>{largest_sku}</b> — <b>{largest_brand}</b> — <b>{largest_dvt}</b> — <b>{largest_val:,.0f} ₫</b> (Tháng {largest_month}).<br><br>
                    • {other_tx_html}
                </div>
                """, unsafe_allow_html=True)
                
                # Hiển thị biểu đồ cột trực quan hoá sản lượng từng SKU tiêu thụ của 1 khách hàng
                st.markdown("<br>", unsafe_allow_html=True)
                sku_shares = cust_tx.groupby('SKU')['Quantity'].sum().reset_index()
                sku_shares = sku_shares.sort_values(by='Quantity', ascending=False)
                
                fig_bar_sku = px.bar(
                    sku_shares, 
                    x='SKU', 
                    y='Quantity',
                    text='Quantity',
                    color='SKU',
                    title=f"Sản lượng tiêu thụ chi tiết từng SKU - Khách hàng {cust}",
                    labels={'Quantity': 'Sản lượng tiêu thụ', 'SKU': 'Mã SKU'},
                    color_discrete_sequence=px.colors.qualitative.Pastel
                )
                fig_bar_sku.update_traces(
                    texttemplate='%{text:,.0f}', 
                    textposition='outside',
                    marker=dict(cornerradius=6),
                    hovertemplate="<b>Mã SKU:</b> %{x}<br><b>Sản lượng tiêu thụ:</b> %{y:,.0f} đơn vị<extra></extra>"
                )
                fig_bar_sku.update_layout(
                    showlegend=False,
                    paper_bgcolor='rgba(0,0,0,0)', 
                    plot_bgcolor='rgba(0,0,0,0)',
                    margin=dict(t=50, b=20, l=10, r=10),
                    font=dict(family="Montserrat", color="#1e293b")
                )
                st.plotly_chart(fig_bar_sku, use_container_width=True)
                
                # Hộp phân tích thông minh dựa trên Bar Chart sản lượng
                total_skus = len(sku_shares)
                total_qty = sku_shares['Quantity'].sum()
                top_sku_row = sku_shares.iloc[0]
                top_sku_id = top_sku_row['SKU']
                top_sku_qty = top_sku_row['Quantity']
                top_sku_pct = (top_sku_qty / total_qty) * 100 if total_qty > 0 else 0
                
                st.markdown(f"""
                <div class="smart-card-info" style="margin-top: 15px;">
                    <b style="color:#1e3a8a; font-size:15px;">💡 PHÂN TÍCH TIÊU THỤ THÔNG MINH (UP-TO-DATE):</b><br><br>
                    • Khách hàng <b>{cust}</b> mua nhiều nhất mã SKU <b>{top_sku_id}</b> với sản lượng đạt <b>{top_sku_qty:,.0f} sản phẩm</b>, chiếm khoảng <b>{top_sku_pct:.1f}%</b> tổng sản lượng tiêu thụ của khách hàng này.<br>
                    • Cơ cấu giỏ hàng gồm <b>{total_skus} mã SKU</b> khác nhau, tổng lượng sản phẩm tiêu thụ đạt <b>{total_qty:,.0f} đơn vị</b>.<br>
                    • Bình quan mỗi mã SKU khách hàng tiêu thụ khoảng <b>{total_qty/total_skus:.1f} sản phẩm</b>. Đề xuất ưu tiên chào bán thêm các chủng loại bổ trợ cho mã bán chạy nhất.
                </div>
                """, unsafe_allow_html=True)

        # --- THUẬT TOÁN SO SÁNH PHÂN TÍCH NHIỀU KHÁCH HÀNG (KHI CHỌN >= 2 KHÁCH HÀNG) ---
        elif len(selected_kh) >= 2:
            compare_tx = df_xb_clean[df_xb_clean['Khach_Hang'].isin(selected_kh)].copy()
            
            if not compare_tx.empty:
                cust_revenue = compare_tx.groupby('Khach_Hang')['Value'].sum().reset_index()
                cust_revenue = cust_revenue.sort_values(by='Value', ascending=False)
                
                # Thiết lập bảng xếp hạng đánh giá chi tiết
                eval_lines = []
                for idx, row in enumerate(cust_revenue.iterrows(), 1):
                    eval_lines.append(f"Top {idx}: Khách hàng <b>{row[1]['Khach_Hang']}</b> đem lại doanh thu lũy kế <b>{row[1]['Value']:,.0f} ₫</b>")
                eval_text = "<br>".join(eval_lines)
                
                # Hiển thị thẻ Đánh giá Đa khách hàng
                st.markdown(f"""
                <div class="smart-card-success">
                    <b style="color:#15803d; font-size:16px;">📊 BẢNG XẾP HẠNG DOANH THU KHÁCH HÀNG:</b><br><br>
                    {eval_text}
                </div>
                """, unsafe_allow_html=True)
                
                # Hiển thị biểu đồ cột so sánh tổng doanh thu của từng khách hàng
                st.markdown("<br>", unsafe_allow_html=True)
                fig_compare = px.bar(
                    cust_revenue, 
                    x='Khach_Hang', 
                    y='Value', 
                    color='Khach_Hang',
                    text='Value',
                    title="So sánh Tổng doanh thu lũy kế giữa các khách hàng được chọn",
                    labels={'Value': 'Tổng doanh thu (VND)', 'Khach_Hang': 'Khách hàng'},
                    color_discrete_sequence=px.colors.qualitative.Set2
                )
                # Tối ưu hóa Tooltip Việt hóa cho biểu đồ cột doanh thu đa khách hàng
                fig_compare.update_traces(
                    texttemplate='%{text:,.0f} ₫', 
                    textposition='outside',
                    hovertemplate="<b>Khách hàng:</b> %{x}<br><b>Tổng doanh thu lũy kế:</b> %{y:,.0f} ₫<extra></extra>"
                )
                fig_compare.update_layout(
                    showlegend=False,
                    paper_bgcolor='rgba(0,0,0,0)', 
                    plot_bgcolor='rgba(0,0,0,0)',
                    font=dict(family="Montserrat", color="#1e293b")
                )
                st.plotly_chart(fig_compare, use_container_width=True)

    # --- TAB 3: RỦI RO HẠN DÙNG ---
    with tab3:
        st.markdown("<h3 style='font-weight: 800;'>🚩 Cảnh báo Hàng Cận/Hết Hạn</h3>", unsafe_allow_html=True)
        risk_df = df[df['Het_HSD_Value'] > 0].copy()
        
        st.metric(label="🚨 GIÁ TRỊ THẤT THOÁT", value=f"{risk_df['Het_HSD_Value'].sum():,.0f} ₫")
        
        if not risk_df.empty:
            fig_risk = px.bar(
                risk_df, 
                x='SKU', 
                y='Het_HSD_Value', 
                color='Hang', 
                text='Het_HSD_Value',
                color_discrete_sequence=px.colors.qualitative.Set3,
                labels={
                    'Het_HSD_Value': 'Giá trị hết hạn',
                    'SKU': 'Mã SKU',
                    'Hang': 'Hãng sản xuất'
                }
            )
            # Tối ưu hóa Tooltip Việt hóa cho biểu đồ rủi ro hạn dùng
            fig_risk.update_traces(
                texttemplate='%{text:,.0f}', 
                textposition='outside', 
                marker=dict(cornerradius=6),
                hovertemplate="<b>Mã SKU:</b> %{x}<br><b>Giá trị hết hạn:</b> %{y:,.0f} ₫<extra></extra>"
            )
            fig_risk.update_layout(
                paper_bgcolor='rgba(0,0,0,0)', 
                plot_bgcolor='rgba(0,0,0,0)',
                font=dict(family="Montserrat", color="#1e293b")
            )
            st.plotly_chart(fig_risk, use_container_width=True)
            
            # Bổ sung ĐVT vào bảng thống kê rủi ro hạn dùng
            display_risk_df = risk_df[['SKU', 'Hang', 'DVT', 'Ton_Kho_SL', 'Het_HSD_Value']].copy()
            display_risk_df.columns = ['Mã SKU', 'Hãng', 'ĐVT', 'Số Lượng Tồn Kho', 'Giá trị thất thoát']
            
            styled_risk = display_risk_df.style.format({
                'Số Lượng Tồn Kho': "{:,.0f}", 
                'Giá trị thất thoát': "{:,.0f} ₫"
            })
            st.dataframe(styled_risk, use_container_width=True)
        else: 
            st.markdown("<div class='smart-card-success'><b style='color:#15803d;'>✅ TRẠNG THÁI AN TOÀN:</b> Không ghi nhận rủi ro cận hạn.</div>", unsafe_allow_html=True)

    # --- TAB 4: PHÂN LOẠI SKU THEO DOANH THU ---
    with tab4:
        st.markdown("<h3 style='font-weight: 800;'>🔥 Top 20 SKU Mang Lại Doanh Thu Cao Nhất</h3>", unsafe_allow_html=True)
        
        top_sku = df.sort_values(by='Xuat_Ban_SL', ascending=False).head(20).sort_values(by='Xuat_Ban_SL', ascending=True)
        
        st.markdown("<h4 style='font-weight: 700; margin-top: 10px;'>1. Sản lượng Xuất bán theo Mã SKU</h4>", unsafe_allow_html=True)
        fig_bar = px.bar(
            top_sku, 
            x='Xuat_Ban_SL', 
            y='SKU', 
            orientation='h', 
            color='SKU', 
            color_discrete_sequence=px.colors.qualitative.Pastel, 
            text='Xuat_Ban_SL',
            labels={'Xuat_Ban_SL': 'Sản lượng', 'SKU': 'Mã SKU'}
        )
        # Tối ưu hóa Tooltip Việt hóa cho biểu đồ cột ngang sản lượng top 20
        fig_bar.update_traces(
            texttemplate='%{text:,.0f}', 
            textposition='outside', 
            marker_line_width=0, 
            marker=dict(cornerradius=8),
            hovertemplate="<b>Mã SKU:</b> %{y}<br><b>Sản lượng bán:</b> %{x:,.0f} đơn vị<extra></extra>"
        )
        fig_bar.update_layout(
            showlegend=False, 
            paper_bgcolor='rgba(0,0,0,0)', 
            plot_bgcolor='rgba(0,0,0,0)', 
            height=500, 
            yaxis=dict(automargin=True),
            font=dict(family="Montserrat", color="#1e293b")
        )
        st.plotly_chart(fig_bar, use_container_width=True)
            
        st.markdown("<h4 style='font-weight: 700; margin-top: 30px;'>2. Tỷ trọng Sản lượng xuất bán Top 20 SKU</h4>", unsafe_allow_html=True)
        fig_pie_sales = px.pie(
            top_sku, 
            values='Xuat_Ban_SL', 
            names='SKU', 
            hole=0.4, 
            color='SKU', 
            color_discrete_sequence=px.colors.qualitative.Pastel
        )
        # Tối ưu hóa Tooltip Việt hóa cho biểu đồ tròn tỷ trọng bán top 20
        fig_pie_sales.update_traces(
            textposition='inside', 
            textinfo='percent',
            hovertemplate="<b>Mã SKU:</b> %{label}<br><b>Sản lượng bán:</b> %{value:,.0f} đơn vị<br><b>Tỷ trọng:</b> %{percent}<extra></extra>"
        )
        fig_pie_sales.update_layout(
            height=600, 
            showlegend=False, 
            paper_bgcolor='rgba(0,0,0,0)', 
            plot_bgcolor='rgba(0,0,0,0)', 
            margin=dict(t=30, b=30, l=10, r=10),
            font=dict(family="Montserrat", color="#1e293b")
        )
        st.plotly_chart(fig_pie_sales, use_container_width=True)

    # --- TAB 5: TRA CỨU CHI TIẾT SKU ---
    with tab5:
        st.markdown("<h3 style='font-weight: 800;'>🔍 Tra cứu chi tiết & Đề xuất AI</h3>", unsafe_allow_html=True)
        selected_sku = st.selectbox("Chọn Mã SKU cần phân tích:", df['SKU'].unique())
        
        if selected_sku:
            sku_data = df[df['SKU'] == selected_sku].iloc[0]
            
            c_desc1, c_desc2, c_desc3 = st.columns(3)
            c_desc1.info(f"**📂 Ngành:** {sku_data['Nganh_Hang']}")
            c_desc2.info(f"**🔬 Chủng loại:** {sku_data['Chung_Loai']}")
            c_desc3.info(f"**🏭 Hãng:** {sku_data['Hang']}")
            
            # --- KIỂM TRA & TRUY VẤN DANH SÁCH KHÁCH HÀNG THỰC TẾ ĐANG MUA SKU NÀY ---
            active_customers = df_xb_clean[df_xb_clean['SKU'] == selected_sku]['Khach_Hang'].dropna().unique().tolist()
            active_customers = sorted([str(c).strip() for c in active_customers if str(c).strip() != ''])
            active_customers_count = len(active_customers)
            
            # Tạo chuỗi danh sách định dạng HTML cho Tooltip hiển thị khi trỏ chuột vào
            if active_customers:
                cust_list_html = "".join([f"• {c}<br>" for c in active_customers])
            else:
                cust_list_html = "Chưa ghi nhận khách hàng"
            
            st.markdown("<br>", unsafe_allow_html=True)
            c1, c2, c3, c4 = st.columns(4)
            # Gán ĐVT chính xác vào metric hiển thị Tồn thực tế
            c1.metric("Tồn thực tế", f"{sku_data['Ton_Kho_SL']:,.0f} {sku_data['DVT']}")
            c2.metric("Bán/ngày", f"{sku_data['Daily_Sales']:.2f}")
            
            # Thẻ KH Active có hỗ trợ Popup Tooltip khi rê chuột
            with c3:
                st.markdown(f"""
                <div class="custom-metric-card">
                    <div class="metric-label">KH ACTIVE</div>
                    <div class="metric-value">{active_customers_count:,}</div>
                    <div class="tooltip-text">
                        <strong style="color: #fb923c; font-size: 13px;">Khách hàng đang mua ({active_customers_count}):</strong><br><br>
                        {cust_list_html}
                    </div>
                </div>
                """, unsafe_allow_html=True)
                
            c4.metric("S2S", f"{sku_data['S2S_Months']:.1f} T")
            
            st.markdown("<h4 style='font-weight: 800; margin-top: 30px; margin-bottom: 20px;'>💡 ĐỀ XUẤT ĐIỀU PHỐI</h4>", unsafe_allow_html=True)
            
            # Bổ sung ĐVT vào nội dung đề xuất đặt mua tự động của AI
            if sku_data['Trang_Thai'] == "🔴 ĐỨT HÀNG": 
                st.markdown(f"<div class='smart-card-error'><b style='color:#b91c1c;'>🚨 BÁO ĐỘNG ĐỨT HÀNG:</b> Tồn hiện tại thấp hơn Lead Time. Mua ngay <b>{sku_data['De_Xuat_Mua']:,.0f} {sku_data['DVT']}</b>. Hạn cuối: <b>{sku_data['Ngay_Dat_Hang_Du_Kien']}</b>.</div>", unsafe_allow_html=True)
            elif sku_data['Trang_Thai'] == "🟡 CẦN NHẬP": 
                st.markdown(f"<div class='smart-card-warning'><b style='color:#b45309;'>⚠️ KẾ HOẠCH NHẬP:</b> Đã chạm ngưỡng ROP. Bổ sung <b>{sku_data['De_Xuat_Mua']:,.0f} {sku_data['DVT']}</b> trước ngày <b>{sku_data['Ngay_Dat_Hang_Du_Kien']}</b>.</div>", unsafe_allow_html=True)
            else: 
                st.markdown(f"<div class='smart-card-success'><b style='color:#15803d;'>🟢 AN TOÀN:</b> Chưa cần nhập thêm. Dự kiến đến <b>{sku_data['Ngay_Dat_Hang_Du_Kien']}</b> mới cần lên đơn.</div>", unsafe_allow_html=True)
                
            if sku_data['S2S_Months'] > 6: 
                st.markdown(f"<div class='smart-card-error'><b style='color:#b91c1c;'>📦 ĐỌNG VỐN:</b> S2S đạt <b>{sku_data['S2S_Months']:.1f} tháng</b>. Rà soát Sale hoặc ngưng nhập.</div>", unsafe_allow_html=True)
            elif sku_data['S2S_Months'] < 1 and sku_data['Daily_Sales'] > 0: 
                st.markdown("<div class='smart-card-warning'><b style='color:#b45309;'>🔥 ÁP LỰC TIÊU THỤ LỚN:</b> Vòng quay kho dưới 1 tháng. Cần nâng DOI.</div>", unsafe_allow_html=True)
                
            if sku_data['Het_HSD_Value'] > 0: 
                st.markdown(f"<div class='smart-card-error'><b style='color:#b91c1c;'>🚩 RỦI RO HSD:</b> Thiệt hại dự kiến <b>{sku_data['Het_HSD_Value']:,.0f} ₫</b>. Áp dụng FEFO ngay.</div>", unsafe_allow_html=True)

except Exception as e:
    st.error(f"Lỗi hệ thống nội bộ: {e}")
